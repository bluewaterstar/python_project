# encoding: utf-8
"""
@file: newcolumnchange.py
@version: 1.0
@author: Atlantis
@time: 2020/6/4 17:39
@DESC: 新大禹新增字段添加脚本（可同时修改ods和dl任务）
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import copy
import openpyxl
import newyu

class EXCEL():
    def __init__(self, file):
        self.file = file
        # 打开一个xlsx文件
        self.wb = openpyxl.load_workbook(self.file)
        print(self.wb.sheetnames)
        # 读取到指定的Sheet页，sheet就变得神奇了，想要的内容都在这里
        sheets = self.wb.get_sheet_names()
        print(sheets)
        self.sheet = sheets[0]
        print(self.sheet)
        self.ws = self.wb[self.sheet]
        print(self.ws)

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


#获取脚本字段的顺序
def column_order(task_code):

    sql = newyu.tasksql(task_code)
    print(sql)

    sql2 = sql.split("\n")
    print(sql2)

    i = 0
    clumns = {}
    for t in sql2:
        # [] 匹配括号中多个或者一个
        print(t)
        aa = t
        cc = aa.lstrip()
        # tt = re.sub(r"['|,]", "", aa).replace("`", '')
        tt = cc.replace("`", "")
        print('-----sql1')
        print(tt.find("COMMENT", 2))
        if (tt.find("COMMENT", 2) != -1 or tt.find("comment", 2) != -1) and tt.find('PARTITIONED BY') == -1:
            #i += 1
            t2 = aa.split()
            print(t2)
            column_str = t2[0]
            print(column_str)
            clumns[i] = column_str.replace('`', '')
            i += 1

        elif tt.find('PARTITIONED BY') == 0:
            print(clumns)
            return clumns
            break
        print(i)


def createSql1(key, value):
    tablename = key
    odstablename = tablename
    dltablename = 'dl_' + tablename[4:]
    database = value['database']
    tablecomment = value['tablecomment']

    # 对各个sql进行生成
    sql2 = '''-- ods创建表\nCREATE EXTERNAL TABLE IF NOT EXISTS {0}\n(\n'''.format(tablename)

    sql4 = '''-- ods脚本\nINSERT OVERWRITE TABLE {0} PARTITION(ds='{{{{yesterday}}}}')\nSELECT\n'''.format(tablename)

    sql6 = '''-- 创建dl视图\nCREATE OR REPLACE VIEW {1}\n\t(\n'''.format(database, dltablename)

    valuecolumn = copy.deepcopy(value)
    # 字段列表
    columnkeys = [i[0] for i in valuecolumn['columns']]
    # 字段类型列表
    columntypes = [i[1] for i in valuecolumn['columns']]

    maxlengthkeys = len(sorted(columnkeys, key=lambda x: len(x))[-1])
    maxlengthtypes = len(sorted(columntypes, key=lambda x: len(x))[-1])
    # 对sql里面字段进行处理
    tmpcolumn1, tmpcolumn2, tmpcolumn3, tmpcolumn4, tmpcolumn5, tmpdata = '', '', '', '', '', 0  # 字段生成组合

    columns1 = ''
    for j in value['columns']:
        # tmpdata += 1
        sql2 = sql2 + '\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + j[1].upper() + ' ' * (
                maxlengthtypes - len(j[1]) + 4) + "COMMENT '" + j[2] + "',\n"

        tmpcolumn1 = tmpcolumn1 + '\t' + j[0] + ',\n'
        tmpcolumn2 = tmpcolumn2 + '\t\t' + j[0] + ',\n'
        tmpcolumn4 = tmpcolumn4 + '\t\t\t' + j[0] + ',\n'
        tmpcolumn5 = tmpcolumn5 + '\t\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + "COMMENT '" + j[
            2] + "',\n"
        if j[3] == 'PRI':
            tmpdata += 1
            if tmpdata == 1:
                columns1 = j[0]
            else:
                columns1 += ',' + j[0]
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(KEYS['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
        else:
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(DATA['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
    sql2 = sql2[
           :-2] + '''\n)\nCOMMENT '{0}'\nPARTITIONED BY (ds STRING COMMENT '分区')ROW FORMAT DELIMITED\nNULL DEFINED AS ""\nSTORED AS orc\nTBLPROPERTIES ("orc.compress"="SNAPPY");\n\n'''.format(
        tablecomment, database, tablename)

    sql4 = sql4 + tmpcolumn1[
                  :-2] + '\nFROM\n\t(\n' + ' ' * 5 + 'SELECT\n' + tmpcolumn2 + '\t\texecute_type,\n\t\tROW_NUMBER()OVER(PARTITION BY {0} ORDER BY execute_time DESC,execute_filename desc,execute_position desc ,execute_index desc) time_rank\n'.format(
        columns1) + '\tFROM\n\t\t(\n\t\t SELECT\n' + tmpcolumn3 + '\t\t\ttype AS execute_type,\n\t\t\texecute_time,\n\t\t\tfilename AS execute_filename,\n\t\t\tposition AS execute_position,\n\t\t\tindex as execute_index' \
           + '\n\t\tFROM\n\t\t\tdl_binlog.stg_binlog_view\n' \
           + "\t\tWHERE\n\t\t\tds = '{{{{yesterday}}}}'\n\t\t\tAND `address`= '{0}'\n\t\t\tAND `schema` = '{1}'\n\t\t\tAND `table`  = '{2}'\n".format(
        value['address'], value['mysqlschema'], value[
            'mysqltablename']) + '\tUNION ALL\n' + '\t\tSELECT' + tmpcolumn4 + "\t\t\t'a' AS execute_type,\n\t\t\tCAST('0000-00-00 00:00:00' as TIMESTAMP) AS execute_time,\n\t\t\t" \
           + "'a' execute_filename,\n\t\t\t1 as execute_position,\n\t\t\t1 execute_index\n" \
           + "\t\tFROM {0}.{1} a\n\t\tWHERE ds = '{{{{yesterday_1}}}}'\n\t\t) a\n\t) b\nWHERE b.time_rank = 1\n\tAND execute_type<>'DELETE';\n\n".format(
        database, tablename)
    sql6 = sql6 + tmpcolumn5 + '\tds' + ' ' * (maxlengthkeys + 2) + "COMMENT '分区'\n\t)\nCOMMENT'{}'".format(
        tablecomment) + '\n\tAS\nSELECT\n' + tmpcolumn1 + '\tds\nFROM\n\t{0}.{1};\n\n'.format(database, tablename)

    return sql2 + sql4, sql6,  odstablename, dltablename

def xstr(s):
    return '' if s is None else str(s)



if __name__ == '__main__':
    # 文件路径


    filepath = r'D:\大搜车\业务资料\数据湖\初始化\table_column_11.xlsx'
    address = 'rm-bp135p76r7y8eo8p3211.mysql.rds.aliyuncs.com:3306'
    # 获取表信息
    tables, table, tablename, num, num1 = {}, {}, '', 0, 0
    ExcelFile = EXCEL(filepath)
    nrows, nclos = ExcelFile.getRowsClosNum()
    for i in range(1, nrows + 1):
        rowdata = ExcelFile.getRowValues(i)
        print(rowdata)
        mysqlschema2 = rowdata[2].replace('-', '_')
        tablename = 'ods_' + mysqlschema2 + '_' + rowdata[1] + '_' + rowdata[5].lower() + '_dd'
        # stg依赖
        stgtablename = 'stg_' + mysqlschema2 + '_' + rowdata[1] + '_db'
        odstableinit = 'ods_' + mysqlschema2 + '_' + rowdata[1] + '_initialization'
        print(tablename)
        if tablename != table.get('tablename'):
            if table.get('tablename'):
                print('------diyi----')
                task_id = newyu.taskCode(table.get('tablename'))

                print(task_id)
                task_id2 = ''
                if task_id != task_id2:
                    task_id2 = task_id
                column_or2 = column_order(task_id)
                num = len(column_or2)
                for i in range(num):
                    column = column_or2[i]
                    print('---------ttttttttt-------------')
                    print(task_id)
                    print(table.get('tablename'))
                    print(tablename)
                    print(column_or1)
                    columns1.append(column_or1[column])

                    print(column_or1[column])
                    del column_or1[column]
                for key in column_or1:
                    columns1.append(column_or1[key])
                    print(column_or1[key])

                print(columns1)
                table['columns'] = columns1
                tables[table['tablename']] = copy.deepcopy(table)

            columns = []
            columns1 = []
            column_or1 = {}
            table['tablename'] = tablename
            table['mysqltablename'] = rowdata[5].lower()
            table['mysqlschema'] = rowdata[2]
            table['database'] = rowdata[0]
            table['address'] = rowdata[10] + ':3306'
            table['address2'] = rowdata[10]
            table['tablecomment'] = xstr(rowdata[4]).replace(';', ' ').replace('"', ' ').replace("'", '')
            table['stgtablename'] = stgtablename
            table['odstableinit'] = odstableinit
        # 字段类型变换
        if rowdata[7].split('(')[0] == 'varchar':
            columntype = 'string'
        elif rowdata[7].split('(')[0] == 'datetime':
            columntype = 'timestamp'
        elif rowdata[7].split('(')[0] in ['decimal', 'double']:
            columntype = rowdata[7]
        else:
            columntype = rowdata[7].split('(')[0]

        columns.append(['`' + rowdata[6] + '`', xstr(columntype),
                        xstr(rowdata[8]).replace(';', ' ').replace('"', ' ').replace("'", ''),
                        rowdata[13]])
        ##顺序
        column_or1[rowdata[6]] = ['`' + rowdata[6] + '`', xstr(columntype),
                                  xstr(rowdata[8]).replace(';', ' ').replace('"', ' ').replace("'", ''),
                                  rowdata[13]]


    task_id = newyu.taskCode(tablename)
    column_or2 = column_order(task_id)
    print("----------aaaaaa------------")
    print(column_or2)

    num = len(column_or2)
    for i in range(num):
        print(i)
        column = column_or2[i]

        print(column)
        columns1.append(column_or1[column])
        del column_or1[column]
    for key in column_or1:
        columns1.append(column_or1[key])
        print(column_or1[key])

    table['columns'] = columns1

    print(table)
    tables[table['tablename']] = copy.deepcopy(table)
    print(tables)
    num1 = 0
    mysqlschema, mysqlschema2, stgtablename = '', '', ''

    new_ods_file = {'dl_retail': '1OwXzLKata', 'dl_tgc': 'loAHbk9n45', 'dl_cyp': 'eaO4nDRAb2',
                    'dl_czy': 'APaueQNMdD', 'dl_finsv': 'oA4CmngA71', 'dl_ins': 'XJsD6ODKOi',
                    'dl_fin': '6rcDeTRJgb', 'dl_pub': '7NsJbTDbvw', 'dl_cheniu': 'xIKOCzdO0u',
                    'dl_risk': '2eOI6lTkja', 'dl_ycgj': '3pMzaO2OS3', 'dl_scm': 'AZcJP2e8Jh',
                    'dl_dfc': '9nM1Va9p4S', 'dl_inrs': 'GZQH3pYjTz', 'dl_jiaxuan': 'oeK0AeSQBF',
                    'dl_outrs': 'YncJBxZ150'}
    new_dl_file = {'dl_retail': 'wtK2ZFPzu0', 'dl_tgc': 'sSKAV77uFA', 'dl_cyp': 'ULALvKzyWI',
               'dl_czy': 'DOOIoZUpZJ', 'dl_finsv': 'drQjxqD6aa', 'dl_ins': 'kSOsykzb27',
               'dl_fin': 'Q3wnDdd8fU', 'dl_pub': 'XKM3pTs5GB', 'dl_cheniu': '1B82qb8Zd1',
               'dl_risk': 'EXQN6SJCyv', 'dl_ycgj': '5lqEN4k77q', 'dl_scm': '8wuqvi2XuK',
               'dl_dfc': 'qAucM2a0XI', 'dl_inrs': 'azaWUPUiXf', 'dl_jiaxuan': 'M6gpu5ucOo',
               'dl_outrs': '3qA1WIvOWr'}


    for key, value in tables.items():
        num1 += 1
        #print(value)
        mysqlschema = value['mysqlschema']
        hive_database = value['database']
        stgtablename = value['stgtablename']
        address2 = value['address2']
        odstableinit = value['odstableinit']

        ###必填参数
        source_key = 'hive_' + hive_database
        ods_folder_id = new_ods_file[hive_database]
        dl_folder_id = new_dl_file[hive_database]

        # print(num1)
        sql = createSql1(key, value)

        sql1 = sql[0]
        print(sql1)
        sql2 = sql[1]
        odstablename = sql[2]
        dltablename = sql[3]

        stg_dep_tasks = 'replace_stg_binlog_view'

        if ods_folder_id == '':
            break
        if mysqlschema != mysqlschema2:
            mysqlschema2 = mysqlschema
            # 初始化
            #init_scr = task_inita.task_inita_ods(mysqlschema, address2, odstableinit)
            # 虚拟节点

            stg_task = newyu.stg_task1(stgtablename)

            ods_parent_task_code = newyu.taskCode(stgtablename)

            newyu.stg_update_task1(ods_parent_task_code, stgtablename)

            newyu.commit_task(ods_parent_task_code)

            stg_sql = 'select 1'
            #commit_post_ods = url.request_parms(stg_sql, conn_id, source_key, stg_dep_tasks, stgtablename, '942')
            #stg_commit = update_url2.request_parms(stgtablename)
            #stg_task_id = getTaskcode.taskCode(stgtablename)
            #stg_task_dep = getTaskcode.dep_tasks(stg_task_id, stg_sql, conn_id, source_key, stg_dep_tasks, stgtablename,
            #                                     '942')
            num1 = 1

        #createFile(sql1, sql2, sql3, mysqlschema, filepath, num1)

        ods_dep_tasks = stgtablename

        print(ods_dep_tasks)
        #新ods任务
        ods_task_id = newyu.taskCode(odstablename)

        ods_task = newyu.update_task1(ods_task_id, ods_parent_task_code, stgtablename, odstablename, ods_folder_id,
                                      sql1, source_key)
        ods_commit = newyu.commit_task(ods_task_id)



        dl_task_id = newyu.taskCode(dltablename)

        dl_task = newyu.update_task1(dl_task_id, ods_task_id, ods_task_id, dltablename, dl_folder_id,
                                     sql2, source_key)
        dl_commit = newyu.commit_task(dl_task_id)

        # commit_post_dl = url.request_parms(sql2, conn_id, source_key, odstablename, dltablename, dl_folder_id)

