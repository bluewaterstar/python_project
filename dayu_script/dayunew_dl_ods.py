# encoding: utf-8
"""
@file: dayunew_dl_ods.py
@version: 1.0
@author: Atlantis
@time: 2020/6/4 17:36
@DESC: 在新大禹上新建ods和dl任务
"""
import sys

sys.stdout.reconfigure(encoding='utf-8')

import copy, os

import openpyxl

#import url
#import update_url2
#import getTaskcode
import task_inita
import newyu


class EXCEL():
    def __init__(self, file):
        self.file = file
        # 打开一个xlsx文件
        self.wb = openpyxl.load_workbook(self.file)
        print(self.wb.sheetnames)
        # 读取到指定的Sheet页，sheet就变得神奇了，想要的内容都在这里
        sheets = self.wb.get_sheet_names()
        print(sheets)
        self.sheet = sheets[0]
        print(self.sheet)
        self.ws = self.wb[self.sheet]
        print(self.ws)

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


def createSql(key, value):
    tablename = key
    tmptablename = 'tmp_ods20200106' + tablename[4:]
    # tmp1tablename = 'tmp1_' + tablename[4:]

    odstablename = tablename
    dltablename = 'dl_' + tablename[4:]
    database = value['database']
    tablecomment = value['tablecomment']

    # 对各个sql进行生成
    # sql1 = '''ALTER VIEW {0}.{1}\nRENAME TO {0}.{2};\n\n'''.format(database, tablename, tmptablename)
    sql2 = '''-- ods创建表\nCREATE TABLE IF NOT EXISTS {0}.{1}\n(\n'''.format(database, tablename)

    sql3 = '''-- \nINSERT OVERWRITE TABLE {0}.{1} PARTITION(ds='{{{{yesterday|delta(-2)}}}}')\nSELECT\n'''.format(
        database, tablename)

    print(sql3)
    sql4 = '''-- ods脚本\nINSERT OVERWRITE TABLE {0}.{1} PARTITION(ds='{{{{yesterday}}}}')\nSELECT\n'''.format(database,
                                                                                                             tablename)
    # sql5 = '''-- 修改dl视图\nALTER VIEW {0}.{1}\nRENAME TO {0}.{2};\n\n'''.format(database, dltablename, tmp1tablename)
    sql6 = '''-- 创建dl视图\nCREATE OR REPLACE VIEW {0}.{1}\n\t(\n'''.format(database, dltablename)

    # sql7 = '''select ds,count(1) from\n{0}.{1}\ngroup by ds;\n\n'''.format(database, tablename)
    # sql8 = '''select `__ds__`,count(1) from\n{0}.{1}\ngroup by `__ds__`;\n\n'''.format(database, tmptablename)
    # sql9 = '''select ds,count(1) from\n{0}.{1}\ngroup by ds;\n\n'''.format(database, dltablename)
    valuecolumn = copy.deepcopy(value)
    # 字段列表
    columnkeys = [i[0] for i in valuecolumn['columns']]
    # 字段类型列表
    columntypes = [i[1] for i in valuecolumn['columns']]

    maxlengthkeys = len(sorted(columnkeys, key=lambda x: len(x))[-1])
    maxlengthtypes = len(sorted(columntypes, key=lambda x: len(x))[-1])
    # 对sql里面字段进行处理
    tmpcolumn1, tmpcolumn2, tmpcolumn3, tmpcolumn4, tmpcolumn5, tmpdata = '', '', '', '', '', 0  # 字段生成组合

    columns1 = ''
    for j in value['columns']:
        # tmpdata += 1
        sql2 = sql2 + '\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + j[1].upper() + ' ' * (
                maxlengthtypes - len(j[1]) + 4) + "COMMENT '" + j[2] + "',\n"

        tmpcolumn1 = tmpcolumn1 + '\t' + j[0] + ',\n'
        tmpcolumn2 = tmpcolumn2 + '\t\t' + j[0] + ',\n'
        tmpcolumn4 = tmpcolumn4 + '\t\t\t' + j[0] + ',\n'
        tmpcolumn5 = tmpcolumn5 + '\t\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + "COMMENT '" + j[
            2] + "',\n"
        if j[3] == 'PRI':
            tmpdata += 1
            if tmpdata == 1:
                columns1 = j[0]
            else:
                columns1 += ',' + j[0]
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(KEYS['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
        else:
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(DATA['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
    sql2 = sql2[
           :-2] + '''\n)\nCOMMENT '{0}'\nPARTITIONED BY (ds STRING COMMENT '分区')ROW FORMAT DELIMITED\nNULL DEFINED AS ""\nSTORED AS orc\nLOCATION 'jfs://hd01-jfs/apps/hive/warehouse/{1}.db/{2}'\nTBLPROPERTIES ("orc.compress"="SNAPPY");\n\n'''.format(
        tablecomment, database, tablename)

    sql3 = sql3 + tmpcolumn1[:-2] + '''\nFROM {0}.{1}\n;\n\n'''.format(
        database, tmptablename)
    sql4 = sql4 + tmpcolumn1[
                  :-2] + '\nFROM\n\t(\n' + ' ' * 5 + 'SELECT\n' + tmpcolumn2 + '\t\texecute_type,\n\t\tROW_NUMBER()OVER(PARTITION BY {0} ORDER BY execute_time DESC,execute_filename desc,execute_position desc ,execute_index desc) time_rank\n'.format(
        columns1) + '\tFROM\n\t\t(\n\t\t SELECT\n' + tmpcolumn3 + '\t\t\ttype AS execute_type,\n\t\t\texecute_time,\n\t\t\tfilename AS execute_filename,\n\t\t\tposition AS execute_position,\n\t\t\tindex as execute_index' \
           + '\n\t\tFROM\n\t\t\tdl_binlog.stg_binlog_view\n' \
           + "\t\tWHERE\n\t\t\tds = '{{{{yesterday}}}}'\n\t\t\tAND `address`= '{0}'\n\t\t\tAND `schema` = '{1}'\n\t\t\tAND `table`  = '{2}'\n".format(
        value['address'], value['mysqlschema'], value[
            'mysqltablename']) + '\tUNION ALL\n' + '\t\tSELECT' + tmpcolumn4 + "\t\t\t'a' AS execute_type,\n\t\t\tCAST('0000-00-00 00:00:00' as TIMESTAMP) AS execute_time,\n\t\t\t" \
           + "'a' execute_filename,\n\t\t\t1 as execute_position,\n\t\t\t1 execute_index\n" \
           + "\t\tFROM {0}.{1} a\n\t\tWHERE ds = '{{{{yesterday|delta(-1)}}}}'\n\t\t) a\n\t) b\nWHERE b.time_rank = 1\n\tAND execute_type<>'DELETE';\n\n".format(
        database, tablename)
    sql6 = sql6 + tmpcolumn5 + '\tds' + ' ' * (maxlengthkeys + 2) + "COMMENT '分区'\n\t)\nCOMMENT'{}'".format(
        tablecomment) + '\n\tAS\nSELECT\n' + tmpcolumn1 + '\tds\nFROM\n\t{0}.{1};\n\n'.format(database, tablename)

    return sql2 + sql4, sql6, sql3, odstablename, dltablename


def createSql1(key, value):
    tablename = key
    odstablename = tablename
    dltablename = 'dl_' + tablename[4:]
    database = value['database']
    tablecomment = value['tablecomment']

    # 对各个sql进行生成
    sql2 = '''-- ods创建表\nCREATE TABLE IF NOT EXISTS {0}\n(\n'''.format(tablename)

    sql4 = '''-- ods脚本\nINSERT OVERWRITE TABLE {0} PARTITION(ds='{{{{yesterday}}}}')\nSELECT\n'''.format(tablename)

    sql6 = '''-- 创建dl视图\nCREATE OR REPLACE VIEW {1}\n\t(\n'''.format(database, dltablename)

    valuecolumn = copy.deepcopy(value)
    # 字段列表
    columnkeys = [i[0] for i in valuecolumn['columns']]
    # 字段类型列表
    columntypes = [i[1] for i in valuecolumn['columns']]

    maxlengthkeys = len(sorted(columnkeys, key=lambda x: len(x))[-1])
    maxlengthtypes = len(sorted(columntypes, key=lambda x: len(x))[-1])
    # 对sql里面字段进行处理
    tmpcolumn1, tmpcolumn2, tmpcolumn3, tmpcolumn4, tmpcolumn5, tmpdata = '', '', '', '', '', 0  # 字段生成组合

    columns1 = ''
    for j in value['columns']:
        # tmpdata += 1
        sql2 = sql2 + '\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + j[1].upper() + ' ' * (
                maxlengthtypes - len(j[1]) + 4) + "COMMENT '" + j[2] + "',\n"

        tmpcolumn1 = tmpcolumn1 + '\t' + j[0] + ',\n'
        tmpcolumn2 = tmpcolumn2 + '\t\t' + j[0] + ',\n'
        tmpcolumn4 = tmpcolumn4 + '\t\t\t' + j[0] + ',\n'
        tmpcolumn5 = tmpcolumn5 + '\t\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + "COMMENT '" + j[
            2] + "',\n"
        if j[3] == 'PRI':
            tmpdata += 1
            if tmpdata == 1:
                columns1 = j[0]
            else:
                columns1 += ',' + j[0]
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(KEYS['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
        else:
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(DATA['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
    sql2 = sql2[
           :-2] + '''\n)\nCOMMENT '{0}'\nPARTITIONED BY (ds STRING COMMENT '分区')ROW FORMAT DELIMITED\nNULL DEFINED AS ""\nSTORED AS orc\nTBLPROPERTIES ("orc.compress"="SNAPPY");\n\n'''.format(
        tablecomment, database, tablename)

    sql4 = sql4 + tmpcolumn1[
                  :-2] + '\nFROM\n\t(\n' + ' ' * 5 + 'SELECT\n' + tmpcolumn2 + '\t\texecute_type,\n\t\tROW_NUMBER()OVER(PARTITION BY {0} ORDER BY execute_time DESC,execute_filename desc,execute_position desc ,execute_index desc) time_rank\n'.format(
        columns1) + '\tFROM\n\t\t(\n\t\t SELECT\n' + tmpcolumn3 + '\t\t\ttype AS execute_type,\n\t\t\texecute_time,\n\t\t\tfilename AS execute_filename,\n\t\t\tposition AS execute_position,\n\t\t\tindex as execute_index' \
           + '\n\t\tFROM\n\t\t\tdl_binlog.stg_binlog_view\n' \
           + "\t\tWHERE\n\t\t\tds = '{{{{yesterday}}}}'\n\t\t\tAND `address`= '{0}'\n\t\t\tAND `schema` = '{1}'\n\t\t\tAND `table`  = '{2}'\n".format(
        value['address'], value['mysqlschema'], value[
            'mysqltablename']) + '\tUNION ALL\n' + '\t\tSELECT\n' + tmpcolumn4 + "\t\t\t'a' AS execute_type,\n\t\t\tCAST('0000-00-00 00:00:00' as TIMESTAMP) AS execute_time,\n\t\t\t" \
           + "'a' execute_filename,\n\t\t\t1 as execute_position,\n\t\t\t1 execute_index\n" \
           + "\t\tFROM {0}.{1} a\n\t\tWHERE ds = '{{{{yesterday_1}}}}'\n\t\t) a\n\t) b\nWHERE b.time_rank = 1\n\tAND execute_type<>'DELETE';\n\n".format(
        database, tablename)
    sql6 = sql6 + tmpcolumn5 + '\tds' + ' ' * (maxlengthkeys + 2) + "COMMENT '分区'\n\t)\nCOMMENT'{}'".format(
        tablecomment) + '\n\tAS\nSELECT\n' + tmpcolumn1 + '\tds\nFROM\n\t{0}.{1};\n\n'.format(database, tablename)

    return sql2 + sql4, sql6, odstablename, dltablename


def xstr(s):
    return '' if s is None else str(s)


def createFile(sql1, sql2, sql3, mysqlschema, filepath, num):
    path = os.path.abspath(os.path.dirname(os.path.dirname(filepath))) + '\\Task'
    if not os.path.exists(path):
        os.makedirs(path)
    print(path)
    num2 = '--' + str(num) + '\n'
    with open(path + '\\' + mysqlschema + '_ods.sql', 'a', encoding='utf-8') as odsfile:
        odsfile.write(num2)
        odsfile.write(sql1)

    with open(path + '\\' + mysqlschema + '_dl.sql', 'a', encoding='utf-8') as dlfile:
        dlfile.write(num2)
        dlfile.write(sql2)

    # with open(path+'\\{}.sql'.format(key.upper()),'w') as sqlfile:
    #    sqlfile.write(sql)


if __name__ == '__main__':
    # 文件路径
    filepath = r'D:\大搜车\业务资料\数据湖\初始化\table_column_11.xlsx'
    address = 'rm-bp135p76r7y8eo8p3211.mysql.rds.aliyuncs.com:3306'
    # 获取表信息
    tables, table, tablename = {}, {}, ''
    ExcelFile = EXCEL(filepath)
    nrows, nclos = ExcelFile.getRowsClosNum()

    aa = ''
    aa1 = ''
    for i in range(1, nrows + 1):
        rowdata = ExcelFile.getRowValues(i)
        print(rowdata)
        mysqlschema2 = rowdata[2].replace('-', '_')
        tablename = 'ods_' + mysqlschema2 + '_' + rowdata[1] + '_' + rowdata[5].lower() + '_dd'

        # stg依赖
        stgtablename = 'stg_' + mysqlschema2 + '_' + rowdata[1] + '_db'
        odstableinit = 'ods_' + mysqlschema2 + '_' + rowdata[1] + '_initialization'
        print(tablename)
        if tablename != table.get('tablename'):
            if table.get('tablename'):
                table['columns'] = columns
                tables[table['tablename']] = copy.deepcopy(table)
            columns = []
            table['tablename'] = tablename
            table['mysqltablename'] = rowdata[5].lower()
            table['mysqlschema'] = rowdata[2]
            table['database'] = rowdata[0]
            table['address'] = rowdata[10] + ':3306'
            table['address2'] = rowdata[10]
            table['tablecomment'] = xstr(rowdata[4]).replace(';', ' ').replace('"', ' ').replace("'", '')
            table['stgtablename'] = stgtablename
            table['odstableinit'] = odstableinit
        # 字段类型变换
        if rowdata[7].split('(')[0] == 'varchar':
            columntype = 'string'
        elif rowdata[7].split('(')[0] == 'datetime':
            columntype = 'timestamp'
        elif rowdata[7].split('(')[0] in ['decimal', 'double']:
            columntype = rowdata[7]
        else:
            columntype = rowdata[7].split('(')[0]

        columns.append(['`' + rowdata[6] + '`', xstr(columntype),
                        xstr(rowdata[8]).replace(';', ' ').replace('"', ' ').replace("'", ''),
                        rowdata[13]])

    table['columns'] = columns

    print(table)
    tables[table['tablename']] = copy.deepcopy(table)
    print(tables)
    num1 = 0
    mysqlschema, mysqlschema2, stgtablename = '', '', ''



    new_ods_file = {'dl_retail': '1OwXzLKata', 'dl_tgc': 'loAHbk9n45', 'dl_cyp': 'eaO4nDRAb2',
                    'dl_czy': 'APaueQNMdD', 'dl_finsv': 'oA4CmngA71', 'dl_ins': 'XJsD6ODKOi',
                    'dl_fin': '6rcDeTRJgb', 'dl_pub': '7NsJbTDbvw', 'dl_cheniu': 'xIKOCzdO0u',
                    'dl_risk': '2eOI6lTkja', 'dl_ycgj': '3pMzaO2OS3', 'dl_scm': 'AZcJP2e8Jh',
                    'dl_dfc': '9nM1Va9p4S', 'dl_inrs': 'GZQH3pYjTz', 'dl_jiaxuan': 'oeK0AeSQBF',
                    'dl_outrs': 'YncJBxZ150'}
    new_dl_file = {'dl_retail': 'wtK2ZFPzu0', 'dl_tgc': 'sSKAV77uFA', 'dl_cyp': 'ULALvKzyWI',
               'dl_czy': 'DOOIoZUpZJ', 'dl_finsv': 'drQjxqD6aa', 'dl_ins': 'kSOsykzb27',
               'dl_fin': 'Q3wnDdd8fU', 'dl_pub': 'XKM3pTs5GB', 'dl_cheniu': '1B82qb8Zd1',
               'dl_risk': 'EXQN6SJCyv', 'dl_ycgj': '5lqEN4k77q', 'dl_scm': '8wuqvi2XuK',
               'dl_dfc': 'qAucM2a0XI', 'dl_inrs': 'azaWUPUiXf', 'dl_jiaxuan': 'M6gpu5ucOo',
               'dl_outrs': '3qA1WIvOWr'}

    for key, value in tables.items():
        num1 += 1
        print(value)
        mysqlschema = value['mysqlschema']
        hive_database = value['database']
        stgtablename = value['stgtablename']
        address2 = value['address2']
        odstableinit = value['odstableinit']

        ###必填参数
        source_key = 'hive_' + hive_database
        new_ods_fold = new_ods_file[hive_database]
        new_dl_fold = new_dl_file[hive_database]

        if new_ods_fold == '':
            break

        # print(num1)
        # sql = createSql(key, value)
        # print(type(sql))

        # sql1 = sql[0]
        # print(sql1)
        # sql2 = sql[1]
        # print(sql2)
        # sql3 = sql[2]
        # odstablename = sql[3]
        # dltablename = sql[4]

        stg_dep_tasks = 'replace_stg_binlog_view'
        if mysqlschema != mysqlschema2:
            mysqlschema2 = mysqlschema
            # 初始化
            init_scr = task_inita.task_inita_ods(mysqlschema, address2, odstableinit)
            # 虚拟节点
            # stg_sql = 'select 1'
            # commit_post_ods = url.request_parms(stg_sql, conn_id, source_key, stg_dep_tasks, stgtablename, '942')
            # stg_commit = update_url2.request_parms(stgtablename)
            # stg_task_id = getTaskcode.taskCode(stgtablename)
            # stg_task_dep = getTaskcode.dep_tasks(stg_task_id, stg_sql, conn_id, source_key, stg_dep_tasks, stgtablename, '942')
            num1 = 1

            stg_task = newyu.stg_task1(stgtablename)
            ods_parent_task_code = newyu.taskCode(stgtablename)
            newyu.stg_update_task1(ods_parent_task_code, stgtablename)
            newyu.commit_task(ods_parent_task_code)

        # createFile(sql1, sql2, sql3, mysqlschema, filepath, num1)

        ods_dep_tasks = stgtablename

        print(ods_dep_tasks)
        # odstablename = 'tmp_aa'
        # dltablename = 'tmp_dl_aa'

        # ods任务
        # commit_post_ods = url.request_parms(sql1, conn_id, source_key, stg_dep_tasks, odstablename, ods_folder_id)
        # ods_commit = update_url2.request_parms(odstablename)
        # ods_task_id = getTaskcode.taskCode(odstablename)
        # ods_task_dep = getTaskcode.dep_tasks(ods_task_id, sql1, conn_id, source_key, ods_dep_tasks, odstablename, ods_folder_id)

        # commit_post_dl = url.request_parms(sql2, conn_id, source_key, stg_dep_tasks, dltablename, dl_folder_id)
        # dl_commit = update_url2.request_parms(dltablename)
        # dl_task_id = getTaskcode.taskCode(dltablename)
        # dl_task_dep = getTaskcode.dep_tasks(dl_task_id, sql2, conn_id, source_key, odstablename, dltablename, dl_folder_id)

        # commit_post_dl = url.request_parms(sql2, conn_id, source_key, odstablename, dltablename, dl_folder_id)

        new_sql = createSql1(key, value)

        new_sql1 = new_sql[0]
        print(new_sql1)
        new_sql2 = new_sql[1]
        print(new_sql2)
        odstablename = new_sql[2]

        dltablename = new_sql[3]

        new_ods_task = newyu.task1(odstablename, new_ods_fold, new_sql1, source_key)
        ods_task_id = newyu.taskCode(odstablename)
        ods_task = newyu.update_task1(ods_task_id, ods_parent_task_code, stgtablename, odstablename, new_ods_fold,
                                      new_sql1, source_key)
        ods_commit = newyu.commit_task(ods_task_id)

        new_dl_task = newyu.task1(dltablename, new_dl_fold, new_sql2, source_key)
        dl_task_id = newyu.taskCode(dltablename)
        print(new_dl_file)
        dl_task = newyu.update_task1(dl_task_id, ods_task_id, odstablename, dltablename, new_dl_fold,
                                     new_sql2, source_key)

        dl_commit = newyu.commit_task(dl_task_id)

