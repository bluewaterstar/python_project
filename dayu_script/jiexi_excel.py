# encoding: utf-8
"""
@file: jiexi_excel.py
@version: 1.0
@author: Atlantis
@time: 2020/5/14 17:55
@DESC: 解析excel文件
"""
import sys
import getTaskcode
sys.stdout.reconfigure(encoding='utf-8')
import requests, json
import re
import copy
import openpyxl


class EXCEL():
    def __init__(self, file):
        self.file = file
        # 打开一个xlsx文件
        self.wb = openpyxl.load_workbook(self.file)
        print(self.wb.sheetnames)
        # 读取到指定的Sheet页，sheet就变得神奇了，想要的内容都在这里
        sheets = self.wb.get_sheet_names()
        print(sheets)
        self.sheet = sheets[0]
        print(self.sheet)
        self.ws = self.wb[self.sheet]
        print(self.ws)

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


def tasksql(task_code):
    #task_code = 'ods_card_credit_ipkm_loan_credit_report_dd'
    dev_url = 'http://dayu.souche-inc.com/api/v2/tasks/{}/search'.format(task_code)
    headers = {
        'Cookie': '_security_token_inc=91584515638708100',
        'Host': 'dayu.souche-inc.com',
        'Content-Type': 'application/json;charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36'
    }

    data = {
        '_security_token_inc': '91584515638708100',
        'isDingDing': 'false',
        'tracknick': '15711314694'
    }

    s = requests.Session()

    r = s.get(url=dev_url, params=data, headers=headers)

    j = json.loads(r.text)

    print(j)

    a = j["data"][0]
    print(a)
    task_id = a["task_id"]

    url2 = 'http://dayu.souche-inc.com/api/v2/tasks/{}'.format(task_id)

    a2 = s.get(url=url2, params=data, headers=headers)

    j1 = json.loads(a2.text)
    print(j1)

    j2 = j1["data"]["config"]["sql"]

    print(j2)
    return j2


#获取脚本字段的顺序
def column_order(task_code):
    #task_code = 'ods_card_credit_ipkm_loan_credit_report_dd'
    sql = tasksql(task_code)

    sql2 = sql.split("\n")
    print(sql2)

    i = 0
    clumns = {}
    for t in sql2:
        # [] 匹配括号中多个或者一个
        print(t)
        aa = t
        cc = aa.lstrip()
        # tt = re.sub(r"['|,]", "", aa).replace("`", '')
        tt = cc.replace("`", "")
        print("_____bbb____")
        print(tt.find("COMMENT", 2))
        if tt.find("COMMENT", 2) != -1 and tt.find('PARTITIONED BY') == -1:
            #i += 1
            t2 = aa.split()
            print(t2)
            clumns[i] = t2[0]
            i += 1

        elif tt.find('PARTITIONED BY') == 0:
            print(clumns)
            return clumns
            break
        print(i)


def createSql(key, value):
    tablename = key
    tmptablename = 'tmp_ods20200106' + tablename[4:]
    # tmp1tablename = 'tmp1_' + tablename[4:]






