# encoding: utf-8
"""
@file: columnChange
@version: 1.0
@author: Atlantis
@time: 2020/5/14 15:43
@DeSC: 库表变更就修改大禹任务脚本
"""
import sys
import getTaskcode
sys.stdout.reconfigure(encoding='utf-8')
import requests, json
import re
import copy
import openpyxl


class EXCEL():
    def __init__(self, file):
        self.file = file
        # 打开一个xlsx文件
        self.wb = openpyxl.load_workbook(self.file)
        print(self.wb.sheetnames)
        # 读取到指定的Sheet页，sheet就变得神奇了，想要的内容都在这里
        sheets = self.wb.get_sheet_names()
        print(sheets)
        self.sheet = sheets[0]
        print(self.sheet)
        self.ws = self.wb[self.sheet]
        print(self.ws)

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


def tasksql(task_code):
    #task_code = 'ods_card_credit_ipkm_loan_credit_report_dd'
    dev_url = 'http://dayu.souche-inc.com/api/v2/tasks/{}/search'.format(task_code)
    headers = {
        'Cookie': '_security_token_inc=91584515638708100',
        'Host': 'dayu.souche-inc.com',
        'Content-Type': 'application/json;charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36'
    }

    data = {
        '_security_token_inc': '91584515638708100',
        'isDingDing': 'false',
        'tracknick': '15711314694'
    }

    s = requests.Session()

    r = s.get(url=dev_url, params=data, headers=headers)

    j = json.loads(r.text)

    print(j)

    a = j["data"][0]
    print(a)
    task_id = a["task_id"]

    url2 = 'http://dayu.souche-inc.com/api/v2/tasks/{}'.format(task_id)

    a2 = s.get(url=url2, params=data, headers=headers)

    j1 = json.loads(a2.text)
    print(j1)

    j2 = j1["data"]["config"]["sql"]

    print(j2)
    return j2


#获取脚本字段的顺序
def column_order(task_code):
    #task_code = 'ods_card_credit_ipkm_loan_credit_report_dd'
    sql = tasksql(task_code)

    sql2 = sql.split("\n")
    print(sql2)

    i = 0
    clumns = {}
    for t in sql2:
        # [] 匹配括号中多个或者一个
        print(t)
        aa = t
        cc = aa.lstrip()
        # tt = re.sub(r"['|,]", "", aa).replace("`", '')
        tt = cc.replace("`", "")
        print("_____bbb____")
        print(tt.find("COMMENT", 2))
        if tt.find("COMMENT", 2) != -1 and tt.find('PARTITIONED BY') == -1:
            #i += 1
            t2 = aa.split()
            print(t2)
            column_str = t2[0]

            clumns[i] = column_str.replace('`','')
            i += 1

        elif tt.find('PARTITIONED BY') == 0:
            print(clumns)
            return clumns
            break
        print(i)


def createSql(key, value):
    tablename = key
    tmptablename = 'tmp_ods20200106' + tablename[4:]
    # tmp1tablename = 'tmp1_' + tablename[4:]

    odstablename = tablename
    dltablename = 'dl_' + tablename[4:]
    database = value['database']
    tablecomment = value['tablecomment']

    # 对各个sql进行生成
    # sql1 = '''ALTER VIEW {0}.{1}\nRENAME TO {0}.{2};\n\n'''.format(database, tablename, tmptablename)
    sql2 = '''-- ods创建表\nCREATE EXTERNAL TABLE IF NOT EXISTS {0}.{1}\n(\n'''.format(database, tablename)
    sql3 = '''-- \nINSERT OVERWRITE TABLE {0}.{1} PARTITION(ds='{{{{yesterday|delta(-2)}}}}')\nSELECT\n'''.format(
        database, tablename)
    print(sql3)
    sql4 = '''-- ods脚本\nINSERT OVERWRITE TABLE {0}.{1} PARTITION(ds='{{{{yesterday}}}}')\nSELECT\n'''.format(database,
                                                                                                             tablename)
    # sql5 = '''-- 修改dl视图\nALTER VIEW {0}.{1}\nRENAME TO {0}.{2};\n\n'''.format(database, dltablename, tmp1tablename)
    sql6 = '''-- 创建dl视图\nCREATE OR REPLACE VIEW {0}.{1}\n\t(\n'''.format(database, dltablename)
    # sql7 = '''select ds,count(1) from\n{0}.{1}\ngroup by ds;\n\n'''.format(database, tablename)
    # sql8 = '''select `__ds__`,count(1) from\n{0}.{1}\ngroup by `__ds__`;\n\n'''.format(database, tmptablename)
    # sql9 = '''select ds,count(1) from\n{0}.{1}\ngroup by ds;\n\n'''.format(database, dltablename)
    valuecolumn = copy.deepcopy(value)
    # 字段列表
    columnkeys = [i[0] for i in valuecolumn['columns']]
    # 字段类型列表
    columntypes = [i[1] for i in valuecolumn['columns']]

    maxlengthkeys = len(sorted(columnkeys, key=lambda x: len(x))[-1])
    maxlengthtypes = len(sorted(columntypes, key=lambda x: len(x))[-1])
    # 对sql里面字段进行处理
    tmpcolumn1, tmpcolumn2, tmpcolumn3, tmpcolumn4, tmpcolumn5, tmpdata = '', '', '', '', '', 0  # 字段生成组合

    columns1 = ''
    for j in value['columns']:
        #tmpdata += 1
        sql2 = sql2 + '\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + j[1].upper() + ' ' * (
                maxlengthtypes - len(j[1]) + 4) + "COMMENT '" + j[2] + "',\n"
        tmpcolumn1 = tmpcolumn1 + '\t' + j[0] + ',\n'
        tmpcolumn2 = tmpcolumn2 + '\t\t' + j[0] + ',\n'
        tmpcolumn4 = tmpcolumn4 + '\t\t\t' + j[0] + ',\n'
        tmpcolumn5 = tmpcolumn5 + '\t\t' + j[0] + ' ' * (maxlengthkeys - len(j[0]) + 4) + "COMMENT '" + j[
            2] + "',\n"
        if j[3] == 'PRI':
            tmpdata += 1
            if tmpdata == 1:
                columns1 = j[0]
            else:
                columns1 += ',' + j[0]
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(KEYS['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
        else:
            tmpcolumn3 = tmpcolumn3 + '\t\t\t' + "CAST(DATA['{0}'] AS {1})".format(j[0].replace('`', ''),
                                                                                   j[1].upper()) + ' ' * (
                                 maxlengthkeys + maxlengthtypes - len(j[0]) - len(j[1]) + 4) + 'AS' + ' ' * 4 + j[
                             0] + ',\n'
    sql2 = sql2[
           :-2] + '''\n)\nCOMMENT '{0}'\nPARTITIONED BY (ds STRING COMMENT '分区')ROW FORMAT DELIMITED\nNULL DEFINED AS ""\nSTORED AS orc\nLOCATION 'jfs://hd01-jfs/apps/hive/warehouse/{1}.db/{2}'\nTBLPROPERTIES ("orc.compress"="SNAPPY");\n\n'''.format(
        tablecomment, database, tablename)
    sql3 = sql3 + tmpcolumn1[:-2] + '''\nFROM {0}.{1}\n;\n\n'''.format(
        database, tmptablename)
    sql4 = sql4 + tmpcolumn1[
                  :-2] + '\nFROM\n\t(\n' + ' ' * 5 + 'SELECT\n' + tmpcolumn2 + '\t\texecute_type,\n\t\tROW_NUMBER()OVER(PARTITION BY {0} ORDER BY execute_time DESC,execute_filename desc,execute_position desc ,execute_index desc) time_rank\n'.format(
        columns1) + '\tFROM\n\t\t(\n\t\t SELECT\n' + tmpcolumn3 + '\t\t\ttype AS execute_type,\n\t\t\texecute_time,\n\t\t\tfilename AS execute_filename,\n\t\t\tposition AS execute_position,\n\t\t\tindex as execute_index' \
           + '\n\t\tFROM\n\t\t\tdl_binlog.stg_binlog_view\n' \
           + "\t\tWHERE\n\t\t\tds = '{{{{yesterday}}}}'\n\t\t\tAND `address`= '{0}'\n\t\t\tAND `schema` = '{1}'\n\t\t\tAND `table`  = '{2}'\n".format(
        value['address'], value['mysqlschema'], value[
            'mysqltablename']) + '\tUNION ALL\n' + '\t\tSELECT' + tmpcolumn4 + "\t\t\t'a' AS execute_type,\n\t\t\tCAST('0000-00-00 00:00:00' as TIMESTAMP) AS execute_time,\n\t\t\t" \
           + "'a' execute_filename,\n\t\t\t1 as execute_position,\n\t\t\t1 execute_index\n" \
           + "\t\tFROM {0}.{1} a\n\t\tWHERE ds = '{{{{yesterday|delta(-1)}}}}'\n\t\t) a\n\t) b\nWHERE b.time_rank = 1\n\tAND execute_type<>'DELETE';\n\n".format(
        database, tablename)
    sql6 = sql6 + tmpcolumn5 + '\tds' + ' ' * (maxlengthkeys + 2) + "COMMENT '分区'\n\t)\nCOMMENT'{}'".format(
        tablecomment) + '\n\tAS\nSELECT\n' + tmpcolumn1 + '\tds\nFROM\n\t{0}.{1};\n\n'.format(database, tablename)

    return sql2 + sql4, sql6, sql3, odstablename, dltablename

def xstr(s):
    return '' if s is None else str(s)



if __name__ == '__main__':
    # 文件路径


    filepath = r'D:\大搜车\业务资料\数据湖\初始化\table_column_11.xlsx'
    address = 'rm-bp135p76r7y8eo8p3211.mysql.rds.aliyuncs.com:3306'
    # 获取表信息
    tables, table, tablename, num, num1 = {}, {}, '', 0, 0
    ExcelFile = EXCEL(filepath)
    nrows, nclos = ExcelFile.getRowsClosNum()
    for i in range(1, nrows + 1):
        rowdata = ExcelFile.getRowValues(i)
        print(rowdata)
        mysqlschema2 = rowdata[2].replace('-','_')
        tablename = 'ods_' + mysqlschema2 + '_' + rowdata[1] + '_' + rowdata[5].lower() + '_dd'

        # stg依赖
        stgtablename = 'stg_' + mysqlschema2 + '_' + rowdata[1] + '_db'
        odstableinit = 'ods_' + mysqlschema2 + '_' + rowdata[1] + '_initialization'
        print(tablename)
        if tablename != table.get('tablename'):
            if table.get('tablename'):
                column_or2 = column_order(table.get('tablename'))

                num = len(column_or2)
                for i in range(num):
                    column = column_or2[i]
                    columns1.append(column_or1[column])
                    del column_or1[column]
                for key in column_or1:
                    columns1.append(column_or1[key])
                    print(column_or1[key])
                table['columns'] = columns1
                tables[table['tablename']] = copy.deepcopy(table)
            columns = []
            columns1 = []
            column_or1 = {}
            table['tablename'] = tablename
            table['mysqltablename'] = rowdata[5].lower()
            table['mysqlschema'] = rowdata[2]
            table['database'] = rowdata[0]
            table['address'] = rowdata[10] + ':3306'
            table['address2'] = rowdata[10]
            table['tablecomment'] = xstr(rowdata[4]).replace(';', ' ').replace('"', ' ').replace("'", '')
            table['stgtablename'] = stgtablename
            table['odstableinit'] = odstableinit
        # 字段类型变换
        if rowdata[7].split('(')[0] == 'varchar':
            columntype = 'string'
        elif rowdata[7].split('(')[0] == 'datetime':
            columntype = 'timestamp'
        elif rowdata[7].split('(')[0] in ['decimal', 'double']:
            columntype = rowdata[7]
        else:
            columntype = rowdata[7].split('(')[0]

        columns.append(['`' + rowdata[6] + '`', xstr(columntype),
                        xstr(rowdata[8]).replace(';', ' ').replace('"', ' ').replace("'", ''),
                        rowdata[13]])
        ##顺序
        column_or1[rowdata[6]] = ['`' + rowdata[6] + '`', xstr(columntype),
                                  xstr(rowdata[8]).replace(';', ' ').replace('"', ' ').replace("'", ''),
                                  rowdata[13]]

    # table['columns'] = columns

    column_or2 = column_order(tablename)
    print("----------aaaaaa------------")
    print(column_or2)

    num = len(column_or2)
    print("_____nnnnn_____")
    print(num)
    for i in range(num):
        print(i)
        column = column_or2[i]

        ###print(column)
        columns1.append(column_or1[column])
        del column_or1[column]
    for key in column_or1:
        columns1.append(column_or1[key])
        print(column_or1[key])

    table['columns'] = columns1

    print(table)
    tables[table['tablename']] = copy.deepcopy(table)
    print(tables)
    num1 = 0
    mysqlschema, mysqlschema2, stgtablename = '', '', ''

    ods_file = {'dl_ycgj': '951', 'dl_dfc': '954', 'dl_pub': '959', 'dl_tgc': '955', 'dl_jiaxuan': '957',
                'dl_cyp': '952', 'dl_fin': '1179', 'dl_retail': '950', 'dl_scm': '1157', 'dl_finsv': '1119',
                'dl_risk': '1343', 'dl_inrs': '1083', 'dl_ins': '953'}
    dl_file = {'dl_ycgj': '961', 'dl_dfc': '964', 'dl_pub': '969', 'dl_tgc': '965', 'dl_jiaxuan': '967',
               'dl_cyp': '962', 'dl_fin': '1200', 'dl_retail': '960', 'dl_scm': '1158', 'dl_finsv': '1121',
               'dl_risk': '1344', 'dl_inrs': '1081', 'dl_ins': '963'}
    hive_conn_id = {'dl_ycgj': '1090', 'dl_dfc': '1138', 'dl_pub': '1131', 'dl_tgc': '1105', 'dl_jiaxuan': '1128',
                    'dl_cyp': '1103', 'dl_fin': '1129', 'dl_retail': '1135', 'dl_scm': '1133', 'dl_finsv': '1134',
                    'dl_risk': '1132', 'dl_inrs': '1140', 'dl_ins': '1104'}

    for key, value in tables.items():
        num1 += 1
        #print(value)
        mysqlschema = value['mysqlschema']
        hive_database = value['database']
        stgtablename = value['stgtablename']
        address2 = value['address2']
        odstableinit = value['odstableinit']

        ###必填参数
        conn_id = hive_conn_id[hive_database]
        source_key = 'hive_' + hive_database
        ods_folder_id = ods_file[hive_database]
        dl_folder_id = dl_file[hive_database]

        if hive_database == '' or conn_id == '':
            break

        # print(num1)
        sql = createSql(key, value)
        #####print(type(sql))

        sql1 = sql[0]
        print(sql1)
        sql2 = sql[1]
        print(sql2)
        sql3 = sql[2]
        odstablename = sql[3]
        dltablename = sql[4]

        stg_dep_tasks = 'replace_stg_binlog_view'
        if mysqlschema != mysqlschema2:
            mysqlschema2 = mysqlschema
            # 初始化
            #init_scr = task_inita.task_inita_ods(mysqlschema, address2, odstableinit)
            # 虚拟节点
            stg_sql = 'select 1'
            #commit_post_ods = url.request_parms(stg_sql, conn_id, source_key, stg_dep_tasks, stgtablename, '942')
            #stg_commit = update_url2.request_parms(stgtablename)
            #stg_task_id = getTaskcode.taskCode(stgtablename)
            #stg_task_dep = getTaskcode.dep_tasks(stg_task_id, stg_sql, conn_id, source_key, stg_dep_tasks, stgtablename,
            #                                     '942')
            num1 = 1

        #createFile(sql1, sql2, sql3, mysqlschema, filepath, num1)
        ods_dep_tasks = stgtablename

        print(ods_dep_tasks)
        # odstablename = 'tmp_aa'
        # dltablename = 'tmp_dl_aa'

        # ods任务
        #commit_post_ods = url.request_parms(sql1, conn_id, source_key, stg_dep_tasks, odstablename, ods_folder_id)
        #ods_commit = update_url2.request_parms(odstablename)
        ods_task_id = getTaskcode.taskCode(odstablename)
        ods_task_dep = getTaskcode.dep_tasks(ods_task_id, sql1, conn_id, source_key, ods_dep_tasks, odstablename,
                                             ods_folder_id)

        #commit_post_dl = url.request_parms(sql2, conn_id, source_key, stg_dep_tasks, dltablename, dl_folder_id)
        #dl_commit = update_url2.request_parms(dltablename)
        dl_task_id = getTaskcode.taskCode(dltablename)
        dl_task_dep = getTaskcode.dep_tasks(dl_task_id, sql2, conn_id, source_key, odstablename, dltablename,
                                            dl_folder_id)

        # commit_post_dl = url.request_parms(sql2, conn_id, source_key, odstablename, dltablename, dl_folder_id)










